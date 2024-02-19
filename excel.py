from openpyxl import load_workbook
import pandas as pd

# custom_sheet_names = ["Sheet1", "Sheet3"]


def excel_dfs(data_stream, header=0, nrows=None, custom_sheet_names=None):
    # 使用openpyxl加载工作簿,lazy_load=True懒加载（无效）,data_only=True带有公式的单元格将返回它们的值而非公式本身，
    wb = load_workbook(data_stream, read_only=True, data_only=True)
    all_sheet_names = wb.sheetnames
    print(all_sheet_names)
    if custom_sheet_names is None:
        # 如果未指定custom_sheet_names，则获取所有工作表名称
        # 获取所有工作表的名称
        sheet_names = all_sheet_names
        print("未指定custom_sheet_names,读取所有工作表名称:", sheet_names)
    else:
        sheet_names = custom_sheet_names
        print("指定读取的工作表名称custom_sheet_names:", custom_sheet_names)
        # 确保传入的sheet_names列表中的所有名称都存在于工作簿中
        for sheet in custom_sheet_names:
            if sheet not in all_sheet_names:
                raise ValueError(f"Sheet name '{sheet}' not found in the workbook.")

    # 修改或处理工作簿（例如删除行）
    # ws = wb["Sheet1"]
    # sheet_name = wb.worksheets[0]

    # 关闭工作簿以节省资源，因为后续的pd.read_excel将会打开文件流
    wb.close()

    # 创建一个字典来存储每个工作表名称及其对应的DataFrame
    dfs = {}

    # 循环读取每个工作表内容，data_stream数据，sheet_name，sheet名，header，从第几行开始，nrows，从第几行结束
    for sheet_name in sheet_names:
        print("循环读取工作表名称:", sheet_name)
        # 将文件对象的指针回到初始位置
        data_stream.seek(0)
        df = pd.read_excel(
            data_stream,
            sheet_name=sheet_name,
            dtype=str,
            header=header,
            nrows=nrows,
            engine="openpyxl",
        )
        # df = df.iloc[:5, :5]
        # 将DataFrame存储在字典中，工作表名称为键
        dfs[sheet_name] = df

    print("excel_dfs()已经将sharepoint文件数据加载到dfs数据字典!")
    return dfs
